from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill, Border, Side


# 单元格颜色
class CellColor:
    green = Color(rgb="00B050")
    green_fill = PatternFill(start_color=green,
                             end_color=green, fill_type='solid')

    yellow = Color(rgb="FFFF00")
    yellow_fill = PatternFill(start_color=yellow,
                              end_color=yellow, fill_type='solid')

    border_style = Side(style="thin", color='000000')
    black_border = Border(left=border_style, right=border_style,
                          top=border_style, bottom=border_style)

    blue = Color(rgb="00B0F0")
    blue_fill = PatternFill(start_color=blue,
                            end_color=blue, fill_type='solid')