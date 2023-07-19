import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from tool.stytle import CellColor

""" 
    total_version_defects_dict 所有批次的缺陷
    total_yield_dur_dict 所有良率与运行时间
    total_ver_list 所有版本信息
    total_batch_ver_path_dict 所有的db3文件地址
    total_vrp_file_dict 所有vrp文件名
    total_ver_nu_of_batch_dict 所有各版本的批次数量信息
"""


def output_excel(total_version_defects_dict, total_yield_dur_dict, total_ver_list,
                 total_batch_ver_path_dict, total_vrp_file_dict, total_ver_nu_of_batch_dict):
    output_path = ".\\output"
    now_time = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
    create_new_folder_name = os.path.join(output_path, now_time)
    output_excel_path = fr"{create_new_folder_name}\test1.xlsx"

    # 将字典所有键存入列表中 存的是数据集名称
    total_batch_names = [i for i in total_version_defects_dict.keys()]
    total_yield_dur_names = [i for i in total_yield_dur_dict.keys()]

    wb = Workbook()
    number_of_total_batch = len(total_version_defects_dict)  # 数据集数量
    number_of_total_yield_dur = len(total_yield_dur_dict)

    # 汇总(首页)表格的单元格修改范围
    # aggregation_sheet_edit_range = {
    #     'yieldCellRange': {'row': len(total_ver_list), 'col': number_of_total_batch},
    #     'durationCellRange': {'row': len(total_ver_list), 'col': number_of_total_yield_dur},
    #     'batchNuCellRange': {'row': 1, 'col': len(total_ver_nu_of_batch_dict)}
    # }

    ws = wb.active
    ws.title = "汇总表格"  # 将第一个改名
    # 根据数据集 创建工作簿
    for i in range(0, number_of_total_batch, 1):
        wb.create_sheet(f'{total_batch_names[i]}')

    # * 塞入数据current_dataset_nu下标从0开始
    for current_dataset_nu in range(0, number_of_total_batch, 1):
        current_batch_name = total_batch_names[current_dataset_nu]  # 遍历的批次号名字
        ws = wb[f'{current_batch_name}']  # 选中对应的工作簿

        if len(total_version_defects_dict[current_batch_name]) == 0:
            continue  # 该批次没有缺陷

        current_batch = total_version_defects_dict[current_batch_name]  # 当前(遍历)数据集列表

        # 开始插入缺陷名称 注意从第二列开始插入
        for col in range(0, len(current_batch), 1):
            edit_col = int(2 + col)
            ws.cell(
                row=1, column=edit_col).value = current_batch[col]['name']
            ws.cell(row=1, column=edit_col).fill = CellColor.green_fill
            ws.cell(row=1, column=edit_col).border = CellColor.black_border

        # 插入 第一列数据集版本号
        insert_index = 1
        for row in range(0, len(total_ver_list), 1):
            edit_row = int(2 + row)  # 从第二行开始操作
            ws.cell(row=edit_row,
                    column=1).value = total_ver_list[row]
            ws.cell(row=edit_row, column=1).fill = CellColor.yellow_fill
            ws.cell(row=edit_row, column=1).border = CellColor.black_border
            insert_index += 1

        # 插入数据 按先第一列然第二列插入 因为一列是一个缺陷
        for index_of_defects in range(0, len(current_batch), 1):
            # for col in range(0, len(current_batch[row]['ver']), 1):
            current_batch_ver_index = 0  # 当前缺陷有多个版本 表示准备插入版本的下标
            for i in range(0, len(total_ver_list), 1):

                edit_row = int(2 + i)  # 行数从第二行开始操作
                edit_col = int(2 + index_of_defects)  # 列从第二列开始操作

                # 避免准备插入版本的下标
                if current_batch_ver_index < len(current_batch[index_of_defects]['ver']):
                    # 特定数据集版本中的字典(数量为1)
                    current_defect_id_ver_dict = current_batch[index_of_defects]['ver'][current_batch_ver_index]
                    current_version_name = list(current_defect_id_ver_dict)[0]

                if total_ver_list[i] == current_version_name:
                    # 特定数据集版本中的字典(数量为1)
                    current_defect_id_ver_dict = current_batch[index_of_defects]['ver'][current_batch_ver_index]
                    ws.cell(row=edit_row, column=edit_col).value = \
                        int(current_defect_id_ver_dict[current_version_name])
                    current_batch_ver_index += 1
                else:
                    ws.cell(row=edit_row, column=edit_col).value = 0

                # print(current_defect_id_ver_dict)

        # 开始画折线图
        line_chart = LineChart()
        line_chart_data = Reference(ws, min_row=1, max_row=ws.max_row,
                                    min_col=2, max_col=ws.max_column)
        line_chart_title = Reference(ws, min_row=2, max_row=ws.max_row,
                                     min_col=1, max_col=1)

        series = Series(line_chart_data)
        series.marker

        # todo from_rows=False 默认以没一列为数据系列
        line_chart.title = f'{current_batch_name}报告'
        line_chart.add_data(line_chart_data, from_rows=False, titles_from_data=True)
        line_chart.set_categories(line_chart_title)
        ws.add_chart(line_chart, f'F{ws.max_row + 10}')

        # 插入对应的vrp文件信息
        edit_col = ws.max_column + 1
        ws.cell(row=1, column=edit_col).value = "vrp文件"
        ws.cell(row=1, column=edit_col).fill = CellColor.blue_fill
        ws.cell(row=1, column=edit_col).border = CellColor.black_border
        # 当前批次所有vrp文件名字  所有vrp格式 {批次号:{版本号1:文件,版本号2:文件}...}
        current_batch_vrp_file_name = [vN for vN in list(total_vrp_file_dict[current_batch_name])]
        # 新增加入对应版本的vrp信息
        insert_index = 0
        for i in range(0, len(total_ver_list), 1):
            edit_row = int(2 + i)  # 从第二行开始操作
            name = current_batch_vrp_file_name[insert_index]
            if total_ver_list[i] == name:
                ws.cell(row=edit_row, column=edit_col).value = total_vrp_file_dict[current_batch_name][name]
                insert_index += 1
                if insert_index == len(current_batch_vrp_file_name):
                    insert_index -= 1  # 因需遍历总vrp列表 总有批次批次没有跑满全部版本
            else:
                ws.cell(row=edit_row, column=edit_col).value = '此版本无跑仿真'
                ws.cell(row=edit_row, column=edit_col).fill = CellColor.red_fill
                ws.cell(row=edit_row, column=edit_col).border = CellColor.black_border

    # * 插入汇总表格数据
    ws = wb["汇总表格"]

    ws.cell(row=1, column=1).value = "良率"
    ws.cell(row=1, column=1).fill = CellColor.blue_fill
    ws.cell(row=1, column=1).border = CellColor.black_border

    # 插入批次信息 从第二列往右插入
    for col in range(0, number_of_total_yield_dur, 1):
        # link = "test1.xlsx"+f"#{total_batch_names[col]}!A1"
        link = fr"Y:\GeRun\SecondaryWire\4#\VTSimulation\{total_batch_names[col]}"
        edit_col = int(2 + col)
        ws.cell(
            row=1, column=edit_col).value = total_yield_dur_names[col]
        ws.cell(row=1, column=edit_col).style = "Hyperlink"
        ws.cell(row=1, column=edit_col).fill = CellColor.green_fill
        ws.cell(row=1, column=edit_col).border = CellColor.black_border
        ws.cell(row=1, column=edit_col).hyperlink = link

    # 插入全部版本信息 从二行向下插入
    for row in range(0, len(total_ver_list), 1):
        edit_row = int(2 + row)  # 从第二行开始操作
        ws.cell(row=edit_row, column=1).value = total_ver_list[row]
        ws.cell(row=edit_row, column=1).fill = CellColor.yellow_fill
        ws.cell(row=edit_row, column=1).border = CellColor.black_border

    # 开始插入每个批次对应的单元格信息 以列向下(版本)然向右(批次)移动
    for col in range(0, number_of_total_yield_dur, 1):
        current_yield_dur = total_yield_dur_dict[total_yield_dur_names[col]]

        insert_index = 0  # 现current_yield_dur中插入的位置 从0开始
        # 这里需要遍历总ver_list让当前版本与其判断 而不是current_yield_dur
        for row in range(0, len(total_ver_list), 1):
            edit_row = int(2 + row)
            edit_col = int(2 + col)

            # 判断该良率版本是否列相同
            if total_ver_list[row] == current_yield_dur[insert_index][0]:
                value = float(current_yield_dur[insert_index][1].strip('%')) / 100
                ws.cell(row=edit_row, column=edit_col).value = value
                # #! 是否必要 增加超链接功能
                ws.cell(row=edit_row, column=edit_col).style = "Hyperlink"
                link = f'{total_batch_ver_path_dict[total_yield_dur_names[col]][insert_index]}'.replace(
                    r'\\192.168.0.11\Data', r'Y:', 1)
                ws.cell(row=edit_row, column=edit_col).hyperlink = link
                insert_index += 1
                if insert_index == len(current_yield_dur):
                    insert_index -= 1  # 因为遍历的总列表 总有批次没有跑满全部版本
            else:
                ws.cell(row=edit_row, column=edit_col).value = '无跑仿真'
                ws.cell(row=edit_row, column=edit_col).fill = CellColor.red_fill
                ws.cell(row=edit_row, column=edit_col).border = CellColor.black_border

    # 开始画折线图
    line_chart2 = LineChart()
    line_chart2_data = Reference(ws, min_row=1, max_row=ws.max_row,
                                 min_col=2, max_col=ws.max_column)
    line_chart2_title = Reference(ws, min_row=2, max_row=ws.max_row,
                                  min_col=1, max_col=1)

    series = Series(line_chart2_data)
    series.marker

    # todo from_rows=False 默认以没一列为数据系列
    line_chart2.title = f'汇总良率报告 单位(%)'
    line_chart2.add_data(line_chart2_data, from_rows=False, titles_from_data=True)
    line_chart2.set_categories(line_chart2_title)
    line_chart2.y_axis.scaling.max = 1  # 设置Y轴  最大值
    ws.add_chart(line_chart2, f'F{ws.max_column + 6}')

    # * 总表 第二部分
    bottom_row = ws.max_row + 2
    ws.cell(row=bottom_row, column=1).value = "耗时(s)"
    ws.cell(row=bottom_row, column=1).fill = CellColor.blue_fill
    ws.cell(row=bottom_row, column=1).border = CellColor.black_border

    # 开始插入批次信息 从第二列开始插入
    for col in range(0, number_of_total_yield_dur, 1):
        # 修改下超链接
        link = "test1.xlsx" + f"#{total_batch_names[col]}!A1"
        edit_col = int(2 + col)
        ws.cell(
            row=bottom_row, column=edit_col).value = total_yield_dur_names[col]
        ws.cell(row=bottom_row, column=edit_col).style = "Hyperlink"
        ws.cell(row=bottom_row, column=edit_col).fill = CellColor.green_fill
        ws.cell(row=bottom_row, column=edit_col).border = CellColor.black_border
        ws.cell(row=bottom_row, column=edit_col).hyperlink = link

    # 开始插入每个版本信息 从二行开始插入
    for row in range(0, len(total_ver_list), 1):
        edit_row = int(bottom_row + row + 1)  # 从底部加1行插入
        ws.cell(row=edit_row,
                column=1).value = total_ver_list[row]
        ws.cell(row=edit_row, column=1).fill = CellColor.yellow_fill
        ws.cell(row=edit_row, column=1).border = CellColor.black_border

    # 开始插入每个批次对应的单元格信息
    for col in range(0, number_of_total_yield_dur, 1):
        current_yield_dur = total_yield_dur_dict[total_yield_dur_names[col]]

        insert_index = 0  # 单个数据集中 插入的良率位置 从0开始
        for row in range(0, len(total_ver_list), 1):
            edit_row = int(bottom_row + row + 1)
            edit_col = int(2 + col)

            # 判断该良率版本是否列相同
            if total_ver_list[row] == current_yield_dur[insert_index][0]:

                ws.cell(row=edit_row,
                        column=edit_col).value = float(current_yield_dur[insert_index][2])
                ws.cell(row=edit_row, column=edit_col).style = "Hyperlink"
                db3_file_name = total_batch_ver_path_dict[total_yield_dur_names[col]][insert_index].split('\\')[-1]
                link = "http://192.168.0.130:8022/db_file_summary/0/" + db3_file_name
                ws.cell(row=edit_row, column=edit_col).hyperlink = link

                insert_index += 1
                if insert_index == len(current_yield_dur):
                    insert_index -= 1  # 因为遍历的总列表 总有批次没有跑满全部版本
            else:
                ws.cell(row=edit_row, column=edit_col).value = '无跑仿真'
                ws.cell(row=edit_row, column=edit_col).fill = CellColor.red_fill
                ws.cell(row=edit_row, column=edit_col).border = CellColor.black_border

    # 开始画折线图
    dur_line_chart = LineChart()
    dur_line_chart_data = Reference(ws, min_row=bottom_row, max_row=ws.max_row,
                                    min_col=2, max_col=ws.max_column)
    dur_line_chart_title = Reference(ws, min_row=int(bottom_row + 1), max_row=ws.max_row,
                                     min_col=1, max_col=1)

    series2 = Series(dur_line_chart_data)
    series2.marker

    # todo from_rows=False 默认以每一列为数据系列
    dur_line_chart.title = f'汇总时间报告 单位(s)'
    dur_line_chart.add_data(dur_line_chart_data, from_rows=False, titles_from_data=True)
    dur_line_chart.set_categories(dur_line_chart_title)
    ws.add_chart(dur_line_chart, f'F{ws.max_column + 34}')

    # * 总表 第三部分 各版本批次号数量信息
    bottom_row = ws.max_row + 2
    ws.cell(row=bottom_row, column=1).value = "各版本批次数量"
    ws.cell(row=bottom_row, column=1).fill = CellColor.blue_fill
    ws.cell(row=bottom_row, column=1).border = CellColor.black_border

    ws.cell(
        row=bottom_row, column=2).value = '批次数量(个)'
    ws.cell(row=bottom_row, column=2).fill = CellColor.green_fill
    ws.cell(row=bottom_row, column=2).border = CellColor.black_border

    # 插入每个版本信息
    for row in range(0, len(total_ver_list), 1):
        edit_row = int(bottom_row + row + 1)  # 从底部加1行插入
        ws.cell(row=edit_row,
                column=1).value = total_ver_list[row]
        ws.cell(row=edit_row, column=1).fill = CellColor.yellow_fill
        ws.cell(row=edit_row, column=1).border = CellColor.black_border

    # 插入每个版本的批次数量
    for index in range(0, len(total_ver_list), 1):
        edit_row = int(bottom_row + index + 1)  # 从底部加1行插入
        ws.cell(row=edit_row,
                column=2).value = total_ver_nu_of_batch_dict[total_ver_list[index]]

    # 画图
    nu_of_batch_chart = LineChart()
    nu_of_batch_chart_data = Reference(ws, min_row=bottom_row, max_row=ws.max_row,
                                       min_col=2, max_col=2)
    nu_of_batch_chart_title = Reference(ws, min_row=int(bottom_row+1), max_row=ws.max_row,
                                        min_col=1, max_col=1)

    series = Series(nu_of_batch_chart_data)
    series.marker

    # todo from_rows=False 默认以没一列为数据系列
    nu_of_batch_chart.title = f'各版本批次数量 单位(个)'
    nu_of_batch_chart.add_data(nu_of_batch_chart_data, from_rows=False, titles_from_data=True)
    nu_of_batch_chart.set_categories(nu_of_batch_chart_title)
    ws.add_chart(nu_of_batch_chart, f'F{ws.max_column + 18}')

    os.makedirs(fr'{create_new_folder_name}')
    wb.save(output_excel_path)
    os.startfile(fr'{create_new_folder_name}')
